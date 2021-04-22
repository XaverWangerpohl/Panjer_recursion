import openpyxl as xl

from distributions import Distribution
from insurance import Insurance

"""this dictionary contains the distributions. To add a distribution,
append the table in Excel, the following dictionary and add a new class in 
distributions.py and add a new elif expression in the constructor of the 
Distribution class"""
d = ("B", "NB", "P", "LOG", "ENB")
distr = {}
for i in range(1, 6):
    distr[i] = d[i - 1]


def main():
    """Get information from panjer_recursion.xlsx"""
    wb = xl.load_workbook("panjer_recursion.xlsx")
    ws1 = wb['Main']

    """Number of claims distribution"""
    n_o_c = Distribution(typ=distr[ws1['E2'].value], value1=ws1['E3'].value, value2=ws1['E4'].value,
                         trunc=ws1['E6'].value)
    """Amount per claim Distribution"""
    a_p_c = Distribution(typ=distr[ws1['F2'].value], value1=ws1['F3'].value, value2=ws1['F4'].value,
                         trunc=ws1['F6'].value)
    """Initialize an instance of the Insurance class """
    ins = Insurance(n_o_c, a_p_c)

    """set the total premium according to the given probability of ruin"""
    premium = ins.get_premium(ws1['I2'].value)

    """scaling factor"""
    sc = ws1['F5'].value

    """Bold Font"""
    b = xl.styles.Font(name='Calibri', size=11, bold=True, color='FF000000')

    """create the evaluation sheet"""
    ws2 = wb.create_sheet(str(ins))

    ws2['A1'].value = "n"
    ws2['A1'].font = b
    ws2['A1'].fill = xl.styles.PatternFill("solid", fgColor="009682")

    ws2['B1'].value = "P(S=n)"
    ws2['B1'].font = b
    ws2['B1'].fill = xl.styles.PatternFill("solid", fgColor="009682")

    """write,  expected value and variance"""
    ws2['K1'].value = str(ws1['I2'].value * 100) + "% Ruin #"
    ws2['K1'].font = b
    ws2.column_dimensions['K'].width = len(str(ws2['K1'].value))
    ws2['K1'].fill = xl.styles.PatternFill("solid", fgColor="009682")
    ws2['K2'].value = sc * premium

    ws2['L1'].value = "Mean"
    ws2['L1'].font = b
    ws2['L1'].fill = xl.styles.PatternFill("solid", fgColor="009682")
    ws2['L2'].value = sc * ins.get_ex_value()

    ws2['M1'].value = "Variance"
    ws2['M1'].font = b
    ws2['M1'].fill = xl.styles.PatternFill("solid", fgColor="009682")
    ws2['M2'].value = sc ** 2 * ins.get_variance()

    """calculte premiums using cantellis inequality"""

    ws2['N1'].value = "Cantelli"
    ws2['N1'].font = b
    ws2['N1'].fill = xl.styles.PatternFill("solid", fgColor="009682")
    ws2['N2'].value = sc * ins.get_cantelli_premium(ws1['I2'].value)

    ws2['O1'].value = "Sc. Factor"
    ws2['O1'].font = b
    ws2['O1'].fill = xl.styles.PatternFill("solid", fgColor="009682")
    ws2['O2'].value = sc

    """write the data for the plot"""
    res = 0
    for n in range(0, premium + 1):
        ws2.cell(column=1, row=n + 2, value=sc * n)
        ws2.cell(column=2, row=n + 2, value=ins.get_prob_s(n))
        res += ins.get_prob_s(n)
    """Plot the distribution"""
    n = xl.chart.Reference(ws2, min_col=1, min_row=2, max_row=premium + 10)
    pn = xl.chart.Reference(ws2, min_col=2, min_row=1, max_row=premium + 10)

    chart = xl.chart.ScatterChart()
    chart.x_axis.title = 'n'
    chart.y_axis.title = 'p'
    chart.legend = None
    series = xl.chart.Series(pn, n, title_from_data=True)

    chart.series.append(series)
    chart.title = "Total Claims Distribution"
    chart.style = 13

    s = chart.series[0]
    s.marker.symbol = "triangle"
    s.marker.graphicalProperties.solidFill = "009682"  # Marker filling
    s.marker.graphicalProperties.line.solidFill = "009682"  # Marker outline
    s.graphicalProperties.line.noFill = True

    ws2.add_chart(chart, 'c1')

    wb.save("panjer_recursion.xlsx")


if __name__ == '__main__':
    main()
